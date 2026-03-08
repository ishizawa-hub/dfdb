#!/usr/bin/env python3
"""
export-excel.py - 抽出データをExcelファイルに出力
=====================================================
PDF原本と照合できる形式で全監督・作品データを出力する。
各行にPDFページ番号・カラム位置を含めることで、
データのズレを目視で確認可能にする。
"""

import json
import sys
import os
import re
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_FILE = BASE_DIR / "extracted_data" / "directors_v8.json"
OUTPUT_FILE = BASE_DIR / "extracted_data" / "directors_v8.xlsx"

# 最大作品数（カラム数を決定）
MAX_WORKS = 10

# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
YEAR_FILLS = {
    "2023-2024": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "2021-2022": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "2020-2021": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
}
CELL_FONT = Font(name="Meiryo", size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def create_excel(data):
    wb = Workbook()

    # === シート1: 監督一覧 ===
    ws = wb.active
    ws.title = "監督一覧"

    # ヘッダー行
    headers = [
        "No.", "PDF年度", "ページ", "カラム",
        "監督名", "ローマ字名", "電話番号", "Mg電話",
        "会社/所属", "メール", "ウェブサイト",
        "プロフィール", "ポートレート", "作品数",
    ]
    # 作品カラム（最大MAX_WORKS件）
    for i in range(1, MAX_WORKS + 1):
        headers.extend([
            f"作品{i} タイトル",
            f"作品{i} 制作体制",
            f"作品{i} 年",
            f"作品{i} サムネ",
        ])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # データ行
    row_num = 2
    dir_num = 0

    for year in ["2023-2024", "2021-2022", "2020-2021"]:
        if year not in data:
            continue

        year_fill = YEAR_FILLS.get(year, PatternFill())

        for director in data[year]:
            dir_num += 1
            works = director.get("works", [])

            row_data = [
                dir_num,
                year,
                director.get("_page", ""),
                director.get("_column", ""),
                director.get("name", ""),
                director.get("nameRomaji", ""),
                director.get("phone", ""),
                director.get("phoneMg", ""),
                director.get("company", ""),
                director.get("email", ""),
                director.get("website", ""),
                director.get("profile", "")[:100] if director.get("profile") else "",
                director.get("portraitPath", ""),
                len(works),
            ]

            # 作品データ
            for i in range(MAX_WORKS):
                if i < len(works):
                    w = works[i]
                    row_data.extend([
                        w.get("title", ""),
                        w.get("agency", ""),
                        w.get("year", ""),
                        w.get("thumbnailPath", ""),
                    ])
                else:
                    row_data.extend(["", "", "", ""])

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.fill = year_fill
                cell.border = THIN_BORDER
                if col_idx <= 4:
                    cell.alignment = Alignment(horizontal="center")

            row_num += 1

    # 列幅調整
    col_widths = {
        1: 5,    # No.
        2: 12,   # PDF年度
        3: 6,    # ページ
        4: 7,    # カラム
        5: 15,   # 監督名
        6: 18,   # ローマ字名
        7: 16,   # 電話番号
        8: 16,   # Mg電話
        9: 25,   # 会社
        10: 25,  # メール
        11: 25,  # ウェブサイト
        12: 30,  # プロフィール
        13: 30,  # ポートレート
        14: 6,   # 作品数
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 作品列の幅
    for i in range(MAX_WORKS):
        base_col = 15 + i * 4
        ws.column_dimensions[get_column_letter(base_col)].width = 35      # タイトル
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 25  # 制作体制
        ws.column_dimensions[get_column_letter(base_col + 2)].width = 6   # 年
        ws.column_dimensions[get_column_letter(base_col + 3)].width = 30  # サムネ

    # フリーズペイン
    ws.freeze_panes = "E2"

    # === シート2: サマリー ===
    ws2 = wb.create_sheet("サマリー")
    ws2.append(["項目", "2023-2024", "2021-2022", "2020-2021", "合計"])

    for year in ["2023-2024", "2021-2022", "2020-2021"]:
        dirs = data.get(year, [])
        n_dirs = len(dirs)
        n_works = sum(len(d.get("works", [])) for d in dirs)
        n_portraits = sum(1 for d in dirs if d.get("portraitPath"))
        n_thumbs = sum(1 for d in dirs for w in d.get("works", []) if w.get("thumbnailPath"))
        n_no_name_kanji = sum(1 for d in dirs if not any(
            '\u3000' <= c <= '\u9fff' or '\u30a0' <= c <= '\u30ff' or '\u3040' <= c <= '\u309f'
            for c in d.get("name", "")))

        ws2.append([f"{year} 監督数", n_dirs])
        ws2.append([f"{year} 作品数", n_works])
        ws2.append([f"{year} ポートレート数", n_portraits])
        ws2.append([f"{year} サムネイル数", n_thumbs])
        ws2.append([f"{year} 漢字名なし", n_no_name_kanji])
        ws2.append(["---", ""])

    total_dirs = sum(len(data.get(y, [])) for y in data)
    total_works = sum(len(d.get("works", [])) for y in data for d in data.get(y, []))
    ws2.append(["合計監督数", total_dirs])
    ws2.append(["合計作品数", total_works])

    # スタイル
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=5):
        for cell in row:
            cell.font = Font(name="Meiryo", size=10)
    ws2["A1"].font = Font(name="Meiryo", size=10, bold=True)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    # === シート3: 品質問題リスト ===
    ws3 = wb.create_sheet("品質チェック")
    ws3.append(["No.", "年度", "ページ", "カラム", "問題タイプ", "現在の値", "備考"])

    issue_num = 0
    for year in data:
        for d in data[year]:
            name = d.get("name", "")

            # 漢字名なし
            has_cjk = any(
                '\u3000' <= c <= '\u9fff' or '\u30a0' <= c <= '\u30ff' or '\u3040' <= c <= '\u309f'
                for c in name)
            if not has_cjk and name:
                issue_num += 1
                ws3.append([issue_num, year, d.get("_page"), d.get("_column"),
                           "漢字名なし", name, f"ローマ字: {d.get('nameRomaji', '')}"])

            # 名前が短すぎる（1-2文字）
            if len(name) <= 2 and name:
                issue_num += 1
                ws3.append([issue_num, year, d.get("_page"), d.get("_column"),
                           "名前短すぎ", name, ""])

            # ポートレートなし
            if not d.get("portraitPath"):
                issue_num += 1
                ws3.append([issue_num, year, d.get("_page"), d.get("_column"),
                           "ポートレートなし", name, ""])

            # 作品なし
            if not d.get("works"):
                issue_num += 1
                ws3.append([issue_num, year, d.get("_page"), d.get("_column"),
                           "作品なし", name, ""])

            # 文字化けの可能性（特殊文字含む）
            if re.search(r'[ゞ＾ゝ]', name) or (len(name) > 6 and name.count('ー') > 2):
                issue_num += 1
                ws3.append([issue_num, year, d.get("_page"), d.get("_column"),
                           "文字化けの可能性", name, ""])

    for row in ws3.iter_rows(min_row=1, max_row=1, max_col=7):
        for cell in row:
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(name="Meiryo", size=10, bold=True)

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 7
    ws3.column_dimensions["E"].width = 18
    ws3.column_dimensions["F"].width = 25
    ws3.column_dimensions["G"].width = 30

    # 保存
    wb.save(str(OUTPUT_FILE))
    print(f"Excel saved: {OUTPUT_FILE}")
    print(f"  Directors: {dir_num}")
    print(f"  Issues: {issue_num}")


if __name__ == "__main__":
    print(f"Loading {INPUT_FILE}...")
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    create_excel(data)
